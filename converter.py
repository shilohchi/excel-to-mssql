# encoding: utf8
import logging
import os
import pymssql
from openpyxl.reader.excel import load_workbook

def convertToNumber(s):
	t = ""
	for x in s:
		if x == u"\uff0e":
			t += "."
		else:
			t += str(int(x))
	return float(t)

class InfoParser(object):
	def __init__(self, name):
		self.name = name
		self.wb = load_workbook(name)
		self.ws = self.wb.worksheets[0]

	def parse(self):
		row, col = 0, 0
		self.basic_info_start = (0, 0)
		self.basic_info_end = (9, 0)

		self.performace_info_start = (9, 0)
		row = 9
		while row < 20 and self.ws.cell(row=row, column=0).value != u"相":
			row += 1
		if row == 20:
			raise StandardError
		else:
			self.performace_info_end = (row, 0)

		self.research_info_start = (row, 0)
		while row < 40 and self.ws.cell(row=row, column=0).value != u"共":
			row += 1
		if row == 40:
			raise StandardError
		else:
			self.research_info_end = (row, 0)

		self.service_info_start = (row, 0)
		self.service_info_end = (row + 6, 0)

		basic_info = self._parse_basic_info(self.basic_info_start,\
				self.basic_info_end)
		performace_info = self._parse_performace_info(self.performace_info_start,\
				self.performace_info_end)
		research_info = self._parse_research_info(self.research_info_start,\
				self.research_info_end)
		service_info = self._parse_service_info(self.service_info_start,\
				self.service_info_end)

		device_info = {
			"basic_info": basic_info,
			"performace_info": performace_info,
			"research_info": research_info,
			"service_info": service_info
		}
		return device_info

	def _parse_basic_info(self, start, end):
		row, col = start
		basic_info = {
			"picpath": "./2012/10/24/" + os.path.split(self.name)[1][:4] + ".png",
			"name_zh": self.ws.cell(row=row, column=col + 1).value,
			"school": self.ws.cell(row=row, column= col + 3).value,
			"id": unicode(self.ws.cell(row=row + 1, column=col + 10).value),
			"name_en": self.ws.cell(row=row + 2, column=col + 10).value,
			"affiliation": self.ws.cell(row=row + 3, column=col + 10).value,
			"place": self.ws.cell(row=row + 4, column=col + 10).value,
			"manager": self.ws.cell(row=row + 5, column=col + 10).value,
			"country": self.ws.cell(row=row + 5, column=col + 12).value,
			"manufacturer": self.ws.cell(row=row + 6, column=col + 10).value,
			"specification": self.ws.cell(row=row + 7, column=col + 10).value,
			"value": convertToNumber(self.ws.cell(row=row + 8, column=col + 10).value[:-2]),
			"metric": u"万元",	
			"date": self.ws.cell(row=row + 8, column=col + 15).value
		}
		return basic_info

	def _parse_performace_info(self, start, end):
		row, col = start
		while row < self.performace_info_end[0]:
			if self.ws.cell(row=row, column=col + 1).value == u"主要功能":
				row_func = row
				break
			row += 1
		else:
			raise StandardError

		t = []
		row = self.performace_info_start[0]
		while row < row_func:
			t.append(self.ws.cell(row=row, column=col + 2).value)
			row += 1
		for i in range(len(t)):
			if t[i] is None:
				t[i] = ""
		tech_info = "".join(t)

		t = []
		row = row_func
		while row < self.performace_info_end[0]:
			t.append(self.ws.cell(row=row, column=col + 2).value)
			row += 1
		for i in range(len(t)):
			if t[i] is None:
				t[i] = ""
		func_info = "".join(t)

		performace_info = {
			"tech_info": tech_info,
			"func_info": func_info
		}
		return performace_info

	def _parse_research_info(self, start, end):
		row, col = start
		items = []
		for r in range(row, self.research_info_end[0] + 1):
			value = self.ws.cell(row=r, column=col + 1).value
			if not value is None and u"主在学专".find(value[:1]) != -1:
				items.append(r)

		out = {}
		t = []
		for r in range(items[0], items[1]):
			value = self.ws.cell(row=r, column=col + 2).value
			if not value is None:
				t.append(value)
			else:
				break
		out["direction"] = "".join(t)
		
		t = []
		for r in range(items[1], items[2]):
			value = self.ws.cell(row=r, column=col + 2).value
			if not value is None:
				t.append(value)
			else:
				break
		out["projects"] = "".join(t)
			
		out["articles"] = self._parse_articles_info(\
				(items[2], col + 1), (items[3], col + 1))
		
		t = []
		for r in range(items[3], self.research_info_end[0]):
			value = self.ws.cell(row=r, column=col + 2).value
			if not value is None:
				t.append(value)
			else:
				break
		out["patent"] = "".join(t)

		return out

	def _parse_articles_info(self, start, end):
		row, col = start
		items = []
		for r in range(row + 2, end[0]):
			value = self.ws.cell(row=r, column=col + 1).value
			if value in [1, 2, 3]:
				items.append(r)
		items.append(end[0])

		out = [None] * (len(items) - 1)
		for i in range(len(items) - 1):
			t = {
				"author": [],
				"title": [],
				"journal": [],
				"pages": []
			}
			for r in range(items[i], items[i + 1]):
				_author = self.ws.cell(row=r, column=col + 2).value
				if not _author is None:
					t["author"].append(_author)
				
				_title = self.ws.cell(row=r, column=col + 4).value
				if not _title is None:
					t["title"].append(_title)
				
				_journal = self.ws.cell(row=r, column=col + 9).value
				if not _journal is None:
					t["journal"].append(_journal)

				_pages = self.ws.cell(row=r, column=col + 15).value
				if not _pages is None:
					t["pages"].append(unicode(_pages))

			out[i] = {
				"id": self.ws.cell(row=items[i], column=col + 1).value,
				"author": "".join(t["author"]),
				"title": "".join(t["title"]),
				"journal": "".join(t["journal"]),
				"year": unicode(self.ws.cell(row=items[i], column=col + 12).value),
				"code": unicode(self.ws.cell(row=items[i], column=col + 13).value),
				"pages": "".join(t["pages"])
			}
		return out

	def _parse_service_info(self, start, end):
		row, col = start
		service_info = {
			"unionout": self.ws.cell(row=row, column=col + 4).value,
			"unionin": self.ws.cell(row=row + 3, column=col + 4).value,
			"name": self.ws.cell(row=row + 4, column=col + 4).value,
			"tel": unicode(self.ws.cell(row=row + 4, column=col + 7).value),
			"email": self.ws.cell(row=row + 4, column=col + 11).value,
			"opentime": self.ws.cell(row=row + 5, column=col + 2).value
		}
		return service_info
	
	def parse_as_text(self):
		info = self.parse()

		basic_info = info["basic_info"]
		basic_info_text = u"\n".join([
			u"基本信息",
			u"========",
				u"\t名称: %s" % basic_info["name_zh"],
				u"\t所属学校: %s" % basic_info["school"],
				u"\t图片url: %s" % basic_info["picpath"],
				u"\t仪器编名称: %s" % basic_info["name_en"],
				u"\t所属校内单位: %s" % basic_info["affiliation"],
				u"\t放置地点: %s" % basic_info["place"],
				u"\t仪器负责人: %s" % basic_info["manager"],
				u"\t制造商国别: %s" % basic_info["country"],
				u"\t制造厂商: %s" % basic_info["manufacturer"],
				u"\t仪器规格: %s" % basic_info["specification"],
				u"\t价值: %s" % basic_info["value"],
				u"\t单位: %s" % basic_info["metric"],
				u"\t购置日期: %s" % basic_info["date"]
		])

		performace_info = info["performace_info"]
		performace_info_text = u"\n".join([
			u"性能信息",
			u"========",
				u"\t主要技术指标: %s" % performace_info["tech_info"],
				u"\t主要功能特色: %s" % performace_info["func_info"]
		])

		research_info = info["research_info"]
		research_info_text = u"\n".join([
			u"科研信息",
			u"========",
				u"\t主要研究方向: %s" % research_info["direction"],
				u"\t在研项目: %s" % research_info["projects"],
				u"\t专利或奖项: %s" % research_info["patent"]
		])

		service_info = info["service_info"]
		service_info_text = u"\n".join([
			u"共享服务信息",
			u"============",
				u"\t联盟内: %s" % service_info["unionin"],
				u"\t联盟外: %s" % service_info["unionout"],
				u"\t联系人: %s" % service_info["name"],
				u"\t联系电话: %s" % service_info["tel"],
				u"\t电子邮件: %s" % service_info["email"],
				u"\t开放时间: %s" % service_info["opentime"]
		])

		articles_info = info["research_info"]["articles"]
		t = [
			u"论刊信息",
			u"========"
		]

		for s in articles_info:
			t.append(u"\t%s\t%s\t%s\t%s\t%s\t%s\t%s" % (
				s["id"],
				s["author"],
				s["title"],
				s["journal"],
				s["year"],
				s["code"],
				s["pages"]
			))

		articles_info_text = u"\n".join(t)

		text = u"\n".join([
			basic_info_text,
			performace_info_text,
			research_info_text,
			service_info_text,
			articles_info_text
		])
		return text

def storeDeviceInfo(conn, info):
	cur = conn.cursor()
	cur.execute("""
		insert into device (
			name, school, picpath, code, englishname, institution, place,
			manager, producecountry, producer, specmodel,
			worth, measureunit, purchasetime,
			ability, chara,
			researchfiled, ownproject, ownpatent,
			unionout, unionin, personname, 
			persontelephone, personemail, opentime
		) values (
			'%s', '%s', '%s', '%s', '%s', '%s', '%s',
			'%s', '%s', '%s', '%s',
			%s, '%s', '%s',
			'%s', '%s',
			'%s', '%s', '%s',
			'%s', '%s', '%s',
			'%s', '%s', '%s'
		)
	""" % ( 
		info["basic_info"]["name_zh"],
		info["basic_info"]["school"],
		info["basic_info"]["picpath"],
		info["basic_info"]["id"],
		info["basic_info"]["name_en"],
		info["basic_info"]["affiliation"],
		info["basic_info"]["place"],
		info["basic_info"]["manager"],
		info["basic_info"]["country"],
		info["basic_info"]["manufacturer"],
		info["basic_info"]["specification"],
		info["basic_info"]["value"],
		info["basic_info"]["metric"],
		info["basic_info"]["date"],
		info["performace_info"]["tech_info"],
		info["performace_info"]["func_info"],
		info["research_info"]["direction"],
		info["research_info"]["projects"],
		info["research_info"]["patent"],
		info["service_info"]["unionout"],
		info["service_info"]["unionin"],
		info["service_info"]["name"],
		info["service_info"]["tel"],
		info["service_info"]["email"],
		info["service_info"]["opentime"]
	))
	conn.commit()

def storeArticlesInfo(cur, info):
	pass

if __name__ == "__main__":
	pass
